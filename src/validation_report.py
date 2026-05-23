import pandas as pd
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── CONFIGURATION ─────────────────────────────────────────
INPUT_FILE  = '../data/healthcare_dataset.csv'
DB_FILE     = '../data/healthcare.db'
OUTPUT_FILE = '../output/validation_report.xlsx'

# ── 1. LOAD CSV ───────────────────────────────────────────
df = pd.read_csv(INPUT_FILE)
print("✅ Data Loaded!")

# ── 2. LOAD INTO SQLite ───────────────────────────────────
conn = sqlite3.connect(DB_FILE)
df.to_sql('patients', conn, if_exists='replace', index=False)
print("✅ Loaded into SQLite!")

# ── 3. SQL QUERIES FOR VALIDATION ─────────────────────────

# Total records
total = pd.read_sql("SELECT COUNT(*) as Total_Records FROM patients", conn)

# Duplicate names
duplicates = pd.read_sql("""
    SELECT Name, COUNT(*) as Count 
    FROM patients 
    GROUP BY Name 
    HAVING COUNT(*) > 1
    ORDER BY Count DESC
""", conn)

# Negative or zero billing
invalid_billing = pd.read_sql("""
    SELECT Name, Age, [Billing Amount]
    FROM patients
    WHERE [Billing Amount] <= 0
""", conn)

# Age outliers (below 0 or above 120)
age_outliers = pd.read_sql("""
    SELECT Name, Age, [Medical Condition]
    FROM patients
    WHERE Age < 0 OR Age > 120
""", conn)

# Gender check (only Male/Female allowed)
invalid_gender = pd.read_sql("""
    SELECT Name, Gender
    FROM patients
    WHERE Gender NOT IN ('Male', 'Female')
""", conn)

# Summary by condition
condition_summary = pd.read_sql("""
    SELECT [Medical Condition],
           COUNT(*) as Patient_Count,
           ROUND(AVG([Billing Amount]), 2) as Avg_Billing,
           ROUND(SUM([Billing Amount]), 2) as Total_Billing
    FROM patients
    GROUP BY [Medical Condition]
    ORDER BY Patient_Count DESC
""", conn)

# Test results breakdown
test_results = pd.read_sql("""
    SELECT [Test Results], COUNT(*) as Count
    FROM patients
    GROUP BY [Test Results]
""", conn)

conn.close()
print("✅ SQL Queries Done!")

# ── 4. STYLES ─────────────────────────────────────────────
header_font  = Font(bold=True, color="FFFFFF", size=11)
header_fill  = PatternFill("solid", fgColor="2E75B6")
red_fill     = PatternFill("solid", fgColor="FF0000")
green_fill   = PatternFill("solid", fgColor="70AD47")
yellow_fill  = PatternFill("solid", fgColor="FFD966")
title_font   = Font(bold=True, size=15, color="2E75B6")

def style_header(ws):
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

def set_col_widths(ws, width=25):
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = width

# ── 5. CREATE EXCEL REPORT ────────────────────────────────
wb = Workbook()

# ── SHEET 1: VALIDATION SUMMARY ───────────────────────────
ws1 = wb.active
ws1.title = "Validation Summary"
ws1['A1'] = "🔍 Healthcare Data Validation Report"
ws1['A1'].font = title_font
ws1.merge_cells('A1:B1')
ws1.append([])
ws1.append(["Generated On", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
ws1.append([])
ws1.append(["Check", "Result"])
for cell in ws1[5]:
    cell.font = header_font
    cell.fill = header_fill

checks = [
    ["Total Records",          total['Total_Records'][0]],
    ["Duplicate Names Found",  len(duplicates)],
    ["Invalid Billing Amount", len(invalid_billing)],
    ["Age Outliers",           len(age_outliers)],
    ["Invalid Gender Values",  len(invalid_gender)],
]
for row in checks:
    ws1.append(row)
    # highlight red if issues found
    last_row = ws1.max_row
    if row[1] > 0 and row[0] != "Total Records":
        ws1.cell(last_row, 2).fill = yellow_fill
    elif row[0] != "Total Records":
        ws1.cell(last_row, 2).fill = green_fill

ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 25

# ── SHEET 2: DUPLICATES ───────────────────────────────────
ws2 = wb.create_sheet("Duplicate Records")
ws2.append(["Name", "Count"])
style_header(ws2)
for row in duplicates.values.tolist():
    ws2.append(row)
set_col_widths(ws2)

# ── SHEET 3: INVALID BILLING ──────────────────────────────
ws3 = wb.create_sheet("Invalid Billing")
ws3.append(["Name", "Age", "Billing Amount"])
style_header(ws3)
for row in invalid_billing.values.tolist():
    ws3.append(row)
    for cell in ws3[ws3.max_row]:
        cell.fill = red_fill
set_col_widths(ws3)

# ── SHEET 4: CONDITION SUMMARY ────────────────────────────
ws4 = wb.create_sheet("Condition Summary")
ws4.append(list(condition_summary.columns))
style_header(ws4)
for row in condition_summary.values.tolist():
    ws4.append(row)
set_col_widths(ws4)

# ── SHEET 5: TEST RESULTS ─────────────────────────────────
ws5 = wb.create_sheet("Test Results Breakdown")
ws5.append(list(test_results.columns))
style_header(ws5)
for row in test_results.values.tolist():
    ws5.append(row)
set_col_widths(ws5)

# ── 6. SAVE ───────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"✅ Validation Report Saved: {OUTPUT_FILE}")
print("🎉 Done! Check the output folder!")